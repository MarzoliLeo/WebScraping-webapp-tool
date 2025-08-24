from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import json
from models import Entity
from scraping.runner import scrape_url
from fastapi.responses import Response, JSONResponse
from io import BytesIO
import pandas as pd

app = FastAPI()
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")



@app.get("/.well-known/appspecific/com.chrome.devtools.json", include_in_schema=False)
def _devtools_probe():
    return Response(status_code=204)

@app.get("/favicon.ico", include_in_schema=False)
def _favicon():
    return Response(status_code=204)

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "result_json": None})

@app.post("/scrape", response_class=HTMLResponse)
async def do_scrape(request: Request,
                    url: str = Form(...),
                    use_browser: bool = Form(False),
                    max_wait_ms: int = Form(2000),
                    respect_robots: bool = Form(True)):
    url_str = str(url)
    items, errors = await scrape_url(url_str, use_browser, max_wait_ms, respect_robots)
    items_json = [Entity(**{**i, "source_url": url_str}).model_dump(mode="json") for i in items]
    payload = {
        "ok": len(items_json) > 0,
        "url": url_str,
        "items": items_json,
        "errors": errors,
    }
    result_json = json.dumps(payload, ensure_ascii=False, indent=2)

    return templates.TemplateResponse("index.html", {
        "request": request,
        "result_json": result_json,  # <-- usa questa chiave
        "form_url": url_str,
        "form_use_browser": "true" if use_browser else "false",
        "form_respect_robots": "true" if respect_robots else "false",
        "form_max_wait_ms": max_wait_ms,
    })


@app.get("/api/scrape")
async def api_scrape(url: str, use_browser: bool = True, max_wait_ms: int = 2000, respect_robots: bool = True):
    url_str = str(url)
    items, errors = await scrape_url(url_str, use_browser, max_wait_ms, respect_robots)
    items_json = [Entity(**{**i, "source_url": url_str}).model_dump(mode="json") for i in items]
    payload = {
        "ok": len(items_json) > 0,
        "url": url_str,
        "items": items_json,
        "errors": errors,
    }
    return JSONResponse(payload)


@app.get("/download.json")
async def download_json(url: str, use_browser: bool = True, max_wait_ms: int = 2000, respect_robots: bool = True):
    url_str = str(url)
    items, errors = await scrape_url(url_str, use_browser, max_wait_ms, respect_robots)
    items_json = [Entity(**{**i, "source_url": url_str}).model_dump(mode="json") for i in items]
    payload = {
        "ok": len(items_json) > 0,
        "url": url_str,
        "items": items_json,
        "errors": errors,
    }
    data = json.dumps(payload, ensure_ascii=False, indent=2)
    return StreamingResponse(iter([data]), media_type="application/json",
                             headers={"Content-Disposition": "attachment; filename=contacts.json"})

@app.get("/download.xlsx")
async def download_xlsx(url: str, use_browser: bool = True, max_wait_ms: int = 2000, respect_robots: bool = True):
    url_str = str(url)
    items, errors = await scrape_url(url_str, use_browser, max_wait_ms, respect_robots)

    def join_list(v): return "\n".join(v) if isinstance(v, list) else (v or "")
    rows = [{
        "Nome": it.get("name") or "",
        "Tipo": it.get("entity_type") or "",
        "Telefono/i": join_list(it.get("phones")),
        "Email": join_list(it.get("emails")),
        "Sito web": it.get("website") or "",
        "Indirizzo": it.get("address") or "",
        "Località": it.get("locality") or "",
        "Regione": it.get("region") or "",
        "CAP": it.get("postal_code") or "",
        "Paese": it.get("country") or "",
        "Rating": it.get("rating") if it.get("rating") is not None else "",
        "URL sorgente": it.get("source_url") or url_str,
    } for it in items]

    df = pd.DataFrame(rows or [{}])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Contatti")
        ws = writer.sheets["Contatti"]
        from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
        thin = Side(border_style="thin", color="FFDDDDDD")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
            cell.border = border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wrap_cols = ["Telefono/i", "Email", "Indirizzo", "Sito web"]
        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max([len(str(v)) if v is not None else 0 for v in df[col]] + [len(col)])
            max_len = min(max_len, 80)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, max_len + 2)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=(col in wrap_cols), vertical="top")
    bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=contacts.xlsx"})
